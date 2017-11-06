#
'''
Created on Aug 12, 2017

@author: Himanshu Ranjan
'''
import openpyxl, smtplib, sys 
import datetime

# open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook('C:\\PythonExFiles\\birthdayData.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
#birthMonth = sheet.cell(row=2, column=lastCol).value 

#print(lastCol)
#print(birthMonth)


# Check each member's birthday status.

birthMembers = {}
birthMembers2 = {}

dt = datetime.datetime.now()
currentDayMon = dt.strftime('%m/%d')

print(currentDayMon)
for r in range (2, sheet.max_row +1):
    birthDate = sheet.cell(row=r, column=lastCol).value
    birthDayMon = birthDate.strftime('%m/%d')
    
    if ((birthDate.strftime('%m/%d') > dt.strftime('%m/%d')) and (birthDate.strftime('%m') == dt.strftime('%m')) and ((int(birthDate.strftime('%d')) - int(dt.strftime('%d'))) < 5)):
        name = sheet.cell(row=r, column=1).value 
        email = sheet.cell(row=r, column=2).value
        birthMembers[name] = email
        print('Upcoming Birthdays:', birthMembers)

    elif (birthDate.strftime('%m/%d') == dt.strftime('%m/%d')):
        name = sheet.cell(row=r, column=1).value 
        email = sheet.cell(row=r, column=2).value
        birthMembers2[name] = email
        print('Today\'s Birthday:', birthMembers2)
    


'''
        
print(unpaidMembers)
 
# Log in to email account

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
mailPass = input('Enter the password to login in to the mail: ')
smtpObj.login('automatepgm@gmail.com', mailPass)

# Send out reminder emails

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" %(latestMonth, name, email)
    print('Seding email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('automatepgm@gmail.com', email , body)
    if sendmailStatus != {}:
        print('There was a problem sending mail to %s: %s' %(email, sendmailStatus) )
    
smtpObj.quit()
'''