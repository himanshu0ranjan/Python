'''
Created on Jul 26, 2017

@author: Himanshu Ranjan
'''
import openpyxl
from excelDemo1 import sheetNames

wb = openpyxl.load_workbook('C:\\PythonExFiles\\example.xlsx')
sheetNames = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetNames[0])
print(sheet)
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)

sheet2 = wb.get_sheet_by_name('Sheet1')
print(sheet2.get_highest_row())
print(sheet2.get_highest_column())