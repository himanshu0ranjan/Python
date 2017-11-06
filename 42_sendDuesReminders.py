'''
Created on Jul 27, 2017

@author: Himanshu Ranjan
'''
import openpyxl, smtplib, sys 

# open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook('C:\\PythonExFiles\\duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value 

print(lastCol)
print(latestMonth)

# Check each member's payment status.

unpaidMembers = {}

for r in range (2, sheet.max_row +1):
    payment = sheet.cell(row=r, column=lastCol).value 
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value 
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
        
print(unpaidMembers)
 
# Log in to email account

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
mailPass = input('Enter the password to login in to the mail: ')
smtpObj.login('youremail@gmail.com', mailPass)

# Send out reminder emails

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" %(latestMonth, name, email)
    print('Seding email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('youremail@gmail.com', email , body)
    if sendmailStatus != {}:
        print('There was a problem sending mail to %s: %s' %(email, sendmailStatus) )
    
smtpObj.quit()
